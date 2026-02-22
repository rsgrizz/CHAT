"""
Created by: Randy Grizzelli
Email: grizzellir@gmail.com
GitHub: https://github.com/rsgrizz
Version: 0.1
Date: 2026-02-21
Purpose: Communication Heuristics Analysis for Triage (CHAT) engine for deterministic prioritization of exported communications.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Generator, List, Optional

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None


@dataclass(frozen=True)
class IngestStats:
    rows_seen: int
    rows_emitted: int


@dataclass(frozen=True)
class IngestRow:
    """
    Raw row from input with provenance.
    data keys are the input headers exactly as read from file.
    source_row is 1 based data row number excluding header.
    """
    data: Dict[str, str]
    source_row: int


def _safe_str(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def sniff_input_type(path: str | Path) -> str:
    """
    Returns: csv or xlsx
    """
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".csv":
        return "csv"
    if ext in {".xlsx", ".xlsm"}:
        return "xlsx"
    raise ValueError(f"Unsupported input type: {ext}")


def iter_csv_rows(
    path: str | Path,
    encoding: str = "utf-8",
    errors: str = "replace",
) -> Generator[IngestRow, None, IngestStats]:
    """
    Streams a CSV and yields IngestRow.

    Determinism:
    - header order is the order in the file
    - source_row is stable and 1 based for the first data row
    """
    p = Path(path)
    rows_seen = 0
    rows_emitted = 0

    with p.open("r", encoding=encoding, errors=errors, newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return IngestStats(rows_seen=0, rows_emitted=0)

        fieldnames = list(reader.fieldnames)

        for idx, row in enumerate(reader, start=1):
            rows_seen += 1
            cleaned: Dict[str, str] = {k: _safe_str(row.get(k)) for k in fieldnames}
            rows_emitted += 1
            yield IngestRow(data=cleaned, source_row=idx)

    return IngestStats(rows_seen=rows_seen, rows_emitted=rows_emitted)


def iter_xlsx_rows(
    path: str | Path,
    sheet_name: Optional[str] = None,
) -> Generator[IngestRow, None, IngestStats]:
    """
    Streams an XLSX using openpyxl read_only mode and yields IngestRow.

    Notes:
    - XLSX is supported for convenience
    - for very large exports CSV is recommended
    """
    if openpyxl is None:
        raise RuntimeError("openpyxl is not available. Install it or convert XLSX to CSV before ingest.")

    p = Path(path)
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)

    try:
        ws = wb[sheet_name] if sheet_name else wb.active

        rows_seen = 0
        rows_emitted = 0

        rows_iter = ws.iter_rows(values_only=True)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            return IngestStats(rows_seen=0, rows_emitted=0)

        headers = [_safe_str(h) for h in header_row]
        if not any(headers):
            return IngestStats(rows_seen=0, rows_emitted=0)

        headers = _dedupe_headers(headers)

        for idx, values in enumerate(rows_iter, start=1):
            rows_seen += 1
            data = {headers[i]: _safe_str(values[i]) if i < len(values) else "" for i in range(len(headers))}
            rows_emitted += 1
            yield IngestRow(data=data, source_row=idx)

        return IngestStats(rows_seen=rows_seen, rows_emitted=rows_emitted)
    finally:
        wb.close()


def _dedupe_headers(headers: List[str]) -> List[str]:
    """
    If an export has duplicate column names, suffix them deterministically.
    Example: Message, Message becomes Message, Message_2
    """
    seen: Dict[str, int] = {}
    out: List[str] = []
    for h in headers:
        base = h if h else "COL"
        count = seen.get(base, 0) + 1
        seen[base] = count
        out.append(base if count == 1 else f"{base}_{count}")
    return out


def iter_rows_auto(
    path: str | Path,
    sheet_name: Optional[str] = None,
) -> Generator[IngestRow, None, IngestStats]:
    """
    Chooses CSV or XLSX based on extension.
    """
    t = sniff_input_type(path)
    if t == "csv":
        return iter_csv_rows(path)
    return iter_xlsx_rows(path, sheet_name=sheet_name)